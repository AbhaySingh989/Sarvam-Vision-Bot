from bot.contracts.comparison import DiffRow, DiffReport
from bot.export.excel import create_comparison_workbook
from openpyxl import load_workbook
from io import BytesIO

def test_create_comparison_workbook():
    report = DiffReport(
        changed_rows=[
            DiffRow(
                header_hierarchy="1. Introduction",
                doc_a_text="Old intro",
                doc_a_page="1",
                doc_b_text="New intro",
                doc_b_page="1",
                what_changed="Modified",
                change_summary_3_points="1. Updated context.\n2. Added detail.\n3. Clarified tone."
            )
        ]
    )

    excel_bytes = create_comparison_workbook(report)
    assert len(excel_bytes) > 0

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb.active
    assert ws.title == "Comparison_Results"
    assert ws.freeze_panes == "A2"

    # Check column widths
    assert ws.column_dimensions["A"].width == 35
    assert ws.column_dimensions["G"].width == 68

    # Check header styling
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.color.rgb == "00FFFFFF"
    assert header_cell.fill.start_color.rgb == "001F4E78"
    assert header_cell.value == "header_hierarchy"
