from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bot.contracts.comparison import DiffReport

def create_comparison_workbook(report: DiffReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison_Results"

    # 1. Sheet title
    # 2. Header style: Bold white font, Solid header fill (1F4E78), Centered
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "header_hierarchy", "doc_a_text", "doc_a_page",
        "doc_b_text", "doc_b_page", "what_changed",
        "change_summary_3_points"
    ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 3. Freeze top row
    ws.freeze_panes = "A2"

    # 4. Constant column widths
    column_widths = {
        "A": 35,
        "B": 60,
        "C": 14,
        "D": 60,
        "E": 14,
        "F": 34,
        "G": 68
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 5. Data cell style: Wrap text, Top vertical alignment
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_data in report.changed_rows:
        ws.append([
            row_data.header_hierarchy,
            row_data.doc_a_text,
            row_data.doc_a_page,
            row_data.doc_b_text,
            row_data.doc_b_page,
            row_data.what_changed,
            row_data.change_summary_3_points
        ])

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = data_align

    # Optional Polish: Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

from bot.contracts.entity import EntityReport

def create_entity_workbook(report: EntityReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Entity_Results"

    # Header style: Bold white font, Solid header fill (1F4E78), Centered
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["entity", "value", "source_snippet", "page_number"]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Freeze top row
    ws.freeze_panes = "A2"

    # Constant column widths
    column_widths = {
        "A": 28, # entity
        "B": 36, # value
        "C": 72, # source_snippet
        "D": 14, # page_number
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Data cell style: Wrap text, Top vertical alignment
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_data in report.extracted_entities:
        ws.append([
            row_data.entity,
            row_data.value,
            row_data.source_snippet,
            row_data.page_number
        ])

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = data_align

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
